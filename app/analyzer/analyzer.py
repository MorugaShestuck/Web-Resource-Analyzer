import hashlib
from app.utils.hash_table import HashTable
from app.parser.ContentParser import ContentParser
import openpyxl


class Analyzer:
    """
    A class for analyzing content based on provided keywords and URL.

    Args:
        keywords (list): A list of tuples where each tuple contains a keyword and its associated topic.
        url (str, optional): The URL of the content to analyze. Defaults to an empty string.

    Attributes:
        keywords (list): The list of keyword-topic tuples for analysis.
        url (str): The URL of the content being analyzed.
        score (dict): A dictionary to store topic scores.
        frequent_keywords (dict): A dictionary to store frequently occurring keywords and their counts.

    Methods:
        set_url(url): Set the URL to analyze.
        analyze_content(content): Analyze the content based on the provided keywords.
        get_score(depth=None): Get the topic scores in descending order. Optionally, limit the number of results by depth.
        get_frequent_keywords(): Get frequently occurring keywords and their counts.
        reset(): Reset the topic scores and frequent keywords.
        get_url(): Get the currently set URL.
        load_keywords(keywords): Update the list of keywords for analysis.
    """

    def __init__(self, keywords, url=""):
        self.keywords = keywords
        self.url = url
        self.score = {}
        self.frequent_keywords = {}

    def set_url(self, url):
        """Set the URL to analyze."""
        self.url = url

    def analyze_content(self, content):
        """
        Analyze the content based on the provided keywords.

        Args:
            content (str): The content to analyze.
        """
        for keyword, topic in self.keywords:
            count = content.lower().count(keyword.lower())
            if topic not in self.score:
                self.score[topic] = 0
            self.score[topic] += count

            if count > 0:
                if keyword not in self.frequent_keywords:
                    self.frequent_keywords[keyword] = 0
                self.frequent_keywords[keyword] += count

    def get_score(self, depth=None):
        """
        Get the topic scores in descending order.

        Args:
            depth (int, optional): Limit the number of results to the specified depth. Defaults to None.

        Returns:
            dict: A dictionary containing topic scores.
        """
        sorted_keywords = sorted(self.score.items(), key=lambda x: x[1], reverse=True)

        if depth is None:
            return {k: v for k, v in sorted_keywords if v > 0}
        else:
            result = {}
            count = 0
            for k, v in sorted_keywords:
                if v > 0:
                    result[k] = v
                    count += 1
                    if count >= depth:
                        break
            return result

    def get_frequent_keywords(self):
        """
        Get frequently occurring keywords and their counts.

        Returns:
            dict: A dictionary containing frequently occurring keywords and their counts.
        """
        sorted_keywords = sorted(self.frequent_keywords.items(), key=lambda x: x[1], reverse=True)
        return {k: v for k, v in sorted_keywords if v > 0}

    def reset(self):
        """Reset the topic scores and frequent keywords."""
        self.score = {}
        self.frequent_keywords = {}

    def get_url(self):
        """
        Get the currently set URL.

        Returns:
            str: The currently set URL.
        """
        return self.url

    def load_keywords(self, keywords):
        """
        Update the list of keywords for analysis.

        Args:
            keywords (list): A list of tuples where each tuple contains a keyword and its associated topic.
        """
        self.keywords = keywords



if __name__ == "__main__":
    keywords = HashTable()
    keywords.load("data.json")
    file_path = '../data/sites.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Лист1']
    column_data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        cell_value = row[0]
        column_data.append(cell_value)
    workbook.close()

    for link in column_data:
        p = ContentParser(url=link)
        a = Analyzer(keywords=keywords)
        p.fetch_content()
        p.parse_content()
        a.analyze_content(p.content)
        print(link)
        print(a.get_score())
        print(a.get_frequent_keywords())
        p.reset()
        a.reset()
        print("\n")
